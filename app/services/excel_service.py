from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def generate_attendance_excel(report_data: list, subject_name: str) -> bytes:
    """Generate a styled attendance Excel report and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Title Row ──
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance Report — {subject_name}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Headers ──
    headers = ["S.No", "Student Name", "Roll Number", "Total Classes", "Present", "Absent", "Percentage (%)"]
    col_widths = [6, 25, 15, 14, 10, 10, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[2].height = 20

    # ── Data Rows ──
    alt_fill = PatternFill(start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")

    for idx, student in enumerate(report_data):
        row_num = idx + 3
        percentage = round((student["present"] / student["total"]) * 100, 2) if student["total"] > 0 else 0.0

        values = [
            idx + 1,
            student["name"],
            student["roll_no"],
            student["total"],
            student["present"],
            student["absent"],
            percentage,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = alt_fill

        # Color-code percentage
        pct_cell = ws.cell(row=row_num, column=7)
        pct_cell.number_format = '0.00"%"'
        if percentage < 75:
            pct_cell.font = Font(color="CC0000", bold=True)
        else:
            pct_cell.font = Font(color="27AE60", bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
